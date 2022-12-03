import csv
from datetime import datetime
import openpyxl
import openpyxl.utils
import openpyxl.styles

import matplotlib.pyplot as pyplot
import numpy

import jinja2
import pathlib
import pdfkit


class Vacancy:
    """Класс для представления вакансии

    Attributes:
        name (str): Название вакансии
        salary_from (int): Нижний порог зарплаты
        salary_to (int): Верхний порог зарплаты
        salary_currency (str): Валюта зарплаты
        salary_average (int): Средняя зарплата
        area_name (str): Город
        year (int): Год публикации
    """
    currency_to_rub = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        """Конструктор объекта вакансий

        :param dict vacancy: Словарь вакансии
        """
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(datetime.strptime(vacancy['published_at'], '%Y-%m-%dT%H:%M:%S%z').year)


class Statistic:
    """Класс для представления статистики

    Attributes:
        salary (dict): Зарплата по годам
        vacancies_number (dict): Количество вакансий по названиям
        salary_of_vacancy_name (dict): Зарплата по вакансиям
        vac_count_of_vacancy_name (dict): Количество вакансий по названию
        salary_city (dict): Вакансии по городам
        vac_city_number (dict): Количество вакансий по городам
        count_of_vacancies (int): Количество вакансий
    """
    def __init__(self):
        """Конструктор класса статистики"""
        self.salary = {}
        self.vacancies_number = {}
        self.salary_of_vacancy_name = {}
        self.vac_count_of_vacancy_name = {}
        self.salary_city = {}
        self.vac_city_number = {}
        self.count_of_vacancies = 0

    def year_sal_dynamics(self, vacancy):
        """Составление динамики зарплаты по годам

        :param Vacancy vacancy: Вакансия
        """
        if vacancy.year not in self.salary:
            self.salary[vacancy.year] = [vacancy.salary_average]
        else: self.salary[vacancy.year].append(vacancy.salary_average)

    def year_vac_dynamics(self, vacancy):
        """Составление динамики вакансий по годам

        :param Vacancy vacancy: Вакансия
        """
        if vacancy.year not in self.vacancies_number:
            self.vacancies_number[vacancy.year] = 1
        else: self.vacancies_number[vacancy.year] += 1

    def curr_vac_dynamics(self, vacancy, vacancy_name):
        """Составление динамики зарплаты по конкретной вакансии по городам

        :param Vacancy vacancy: Вакансия
        :param str vacancy_name: Название вакансии
        """
        if vacancy.name.find(vacancy_name) != -1:
            if vacancy.year not in self.salary_of_vacancy_name:
                self.salary_of_vacancy_name[vacancy.year] = [vacancy.salary_average]
            else: self.salary_of_vacancy_name[vacancy.year].append(vacancy.salary_average)

            if vacancy.year not in self.vac_count_of_vacancy_name:
                self.vac_count_of_vacancy_name[vacancy.year] = 1
            else: self.vac_count_of_vacancy_name[vacancy.year] += 1

    def city_sal_dynamics(self, vacancy):
        """Составление динамики зарплаты по городам

        :param Vacancy vacancy: Вакансия
        """
        if vacancy.area_name not in self.salary_city:
            self.salary_city[vacancy.area_name] = [vacancy.salary_average]
        else: self.salary_city[vacancy.area_name].append(vacancy.salary_average)

    def city_count_dynamics(self, vacancy):
        """Составление динамики количества вакансий по городам

        :param Vacancy vacancy: Вакансия
        """
        if vacancy.area_name not in self.vac_city_number:
            self.vac_city_number[vacancy.area_name] = 1
        else: self.vac_city_number[vacancy.area_name] += 1

    def write(self, vacancy, vacancy_name):
        """Заполение статистики

        :param Vacancy vacancy: Вакансия
        :param str vacancy_name: Название определенной вакансии
        """
        self.year_sal_dynamics(vacancy)
        self.year_vac_dynamics(vacancy)
        self.curr_vac_dynamics(vacancy, vacancy_name)
        self.city_sal_dynamics(vacancy)
        self.city_count_dynamics(vacancy)
        self.count_of_vacancies += 1

    def get_stat1(self):
        """ Получить динамику уровня зарплат по годам

        :return dict: Динамика уровня зарплат по годам
        """
        result = {}
        for year, sal in self.salary.items():
            result[year] = int(sum(sal) / len(sal))
        return result

    def get_stat2(self):
        """Получить динамику количества вакансий по годам

        :return dict: Динамика количества вакансий по годам
        """
        return self.vacancies_number

    def get_stat3(self):
        """Получить динамику уровня зарплат по годам для выбранной профессии

        :return dict: Динамика уровня зарплат по годам для выбранной профессии
        """
        if not self.salary_of_vacancy_name:
            self.salary_of_vacancy_name = self.salary.copy()
            self.salary_of_vacancy_name = dict([(key, []) for key, value in self.salary_of_vacancy_name.items()])
        result = {}
        for year, sal in self.salary_of_vacancy_name.items():
            if len(sal) == 0: result[year] = 0
            else: result[year] = int(sum(sal) / len(sal))
        return result

    def get_stat4(self):
        """Получить динамику количества вакансий по годам для выбранной профессии

        :return dict: Динамика количества вакансий по годам для выбранной профессии
        """
        if not self.vac_count_of_vacancy_name:
            self.vac_count_of_vacancy_name = self.vacancies_number.copy()
            self.vac_count_of_vacancy_name = dict(
                [(key, 0) for key, value in self.vac_count_of_vacancy_name.items()])
        return self.vac_count_of_vacancy_name

    def get_stat5and6(self):
        """Получить уровень и долю зарплат по городам (в порядке убывания)

        :return: Уровень и доля зарплат по городам (в порядке убывания)
        """
        result1 = {}
        for city, count in self.vac_city_number.items():
            result1[city] = round(count / self.count_of_vacancies, 4)
        result1 = list(filter(lambda a: a[-1] >= 0.01,
                             [(key, value) for key, value in result1.items()]))
        result1.sort(key=lambda a: a[-1], reverse=True)

        result2 = {}
        for city, sal in self.salary_city.items():
            result2[city] = int(sum(sal) / len(sal))
        result2 = list(filter(lambda a: a[0] in list(dict(result1).keys()),
                             [(key, value) for key, value in result2.items()]))
        result2.sort(key=lambda a: a[-1], reverse=True)

        return dict(result2[:10]), dict(result1[:10])

    def print_statistics(self):
        """Вывести статистику в консоль"""
        print('Динамика уровня зарплат по годам: ' + str(self.get_stat1()))
        print('Динамика количества вакансий по годам: ' + str(self.get_stat2()))
        print('Динамика уровня зарплат по годам для выбранной профессии: ' + str(self.get_stat3()))
        print('Динамика количества вакансий по годам для выбранной профессии: ' + str(self.get_stat4()))
        stat5, stat6 = self.get_stat5and6()
        print('Уровень зарплат по городам (в порядке убывания): ' + str(stat5))
        print('Доля вакансий по городам (в порядке убывания): ' + str(stat6))


class DataSet:
    """Дата-сет для работы с таблицей
    Attributes:
        file_name (str): Название файла
        vacancy_name (str): Название необходимой вакансии
    """
    def __init__(self, file_name, vacancy_name):
        """Конструктор класса DataSet

        :param str file_name: Название файла
        :param str vacancy_name: Название необходимой вакансии
        """
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    def csv_reader(self):
        """Читает CSV файл"""
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    def get_statistic(self):
        """Получить статистические данные

        :return Statistics: Статистика
        """
        statistics = Statistic()

        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            statistics.write(vacancy, self.vacancy_name)

        return statistics


class InputConnect:
    """Начальная точка программы. Объединяет всю логику программы

    Attributes:
        file_name (str): Название файла
        vacancy_name (list): Название необходимой вакансии
    """
    def __init__(self):
        """
        Начало работы программы
        """
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        stats = dataset.get_statistic()
        stats.print_statistics()

        stat5, stat6 = stats.get_stat5and6()
        report = Report(self.vacancy_name, stats.get_stat1(), stats.get_stat2(),
                        stats.get_stat3(), stats.get_stat4(), stat5, stat6)

        report.generate_excel('report.xlsx')
        report.generate_img('graph.png')
        report.generate_pdf('report.pdf')


class Report:
    """Класс для построения статистических графиков

    Attributes:
        wb (Workbook): Вспомогательная переменная для openpyxl
        vacancy_name (str): Название вакансии
        img_filename (str): Название файла для выходного изображения статистики
        stats1 (dict): Динамика уровня зарплат по годам
        stats2 (dict): Динамика количества вакансий по годам
        stats3 (dict): Динамика уровня зарплат по годам для выбранной профессии
        stats4 (dict): Динамика количества вакансий по годам для выбранной профессии
        stats5 (dict): Уровень зарплат по городам (в порядке убывания)
        stats6 (dict): Доля зарплат по городам (в порядке убывания)
    """
    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        """Конструктор класса для построения репортов

        :param str vacancy_name: Название вакансии
        :param dict stats1: Динамика уровня зарплат по годам
        :param dict stats2: Динамика количества вакансий по годам
        :param dict stats3: Динамика уровня зарплат по годам для выбранной профессии
        :param dict stats4: Динамика количества вакансий по годам для выбранной профессии
        :param dict stats5: Уровень зарплат по городам (в порядке убывания)
        :param dict stats6: Доля зарплат по городам (в порядке убывания)
        """
        self.wb = openpyxl.Workbook()
        self.vacancy_name = vacancy_name
        self.img_filename = ''
        self.stats1 = stats1
        self.stats2 = stats2
        self.stats3 = stats3
        self.stats4 = stats4
        self.stats5 = stats5
        self.stats6 = stats6

    def generate_excel(self, filename):
        """Генерирует Excel таблицу

        :param str filename: Название выходного файла
        """
        # Лист "Статистика по годам"
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats1.keys():
            ws1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]: column_widths[i] = len(cell)
                else: column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width + 2

        # Лист "Статистика по городам"
        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws2.append(row)

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]: column_widths[i] = len(cell)
                else: column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width + 2

        font_bold = openpyxl.styles.Font(bold=True)
        for col in 'ABCDE':
            ws1[col + '1'].font = font_bold
            ws2[col + '1'].font = font_bold

        for index, _ in enumerate(self.stats5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = openpyxl.styles.Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                ws2[col + str(row + 1)].border = openpyxl.styles.Border(left=thin, bottom=thin, right=thin, top=thin)

        self.stats1[1] = 1
        for row, _ in enumerate(self.stats1):
            for col in 'ABCDE':
                ws1[col + str(row + 1)].border = openpyxl.styles.Border(left=thin, bottom=thin, right=thin, top=thin)

        self.wb.save(filename)

    def generate_img(self, filename):
        """Генерирует диаграмму со статистикой

        :param str filename: Название выходного файла
        """
        self.img_filename = filename

        fig, ((ax1, ax2), (ax3, ax4)) = pyplot.subplots(nrows=2, ncols=2)

        # Костыль, не знаю почему не работает
        fix_stats1_keys = list(self.stats1.keys())[:-1]
        fix_stats1_value = list(self.stats1.values())[:-1]

        bar1 = ax1.bar(numpy.array(fix_stats1_keys) - 0.4, fix_stats1_value, width=0.4)
        bar2 = ax1.bar(numpy.array(fix_stats1_keys), self.stats3.values(), width=0.4)
        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax1.grid(axis='y')
        ax1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        ax1.set_xticks(numpy.array(fix_stats1_keys) - 0.2, fix_stats1_keys, rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)

        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = ax2.bar(numpy.array(list(self.stats2.keys())) - 0.4, self.stats2.values(), width=0.4)
        bar2 = ax2.bar(numpy.array(list(self.stats2.keys())), self.stats4.values(), width=0.4)
        ax2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
                   prop={'size': 8})
        ax2.set_xticks(numpy.array(list(self.stats2.keys())) - 0.2, list(self.stats2.keys()), rotation=90)
        ax2.grid(axis='y')
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)

        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.stats5.keys()))]),
                 list(reversed(list(self.stats5.values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')

        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.stats6.values()])
        ax4.pie(list(self.stats6.values()) + [other],
                labels=list(self.stats6.keys()) + ['Другие'],
                textprops={'fontsize': 6})

        pyplot.tight_layout()
        pyplot.savefig(filename)

    def generate_pdf(self, filename):
        """Генерирует PDF файл со статистикой

        :param str filename: Название выходного файла
        """
        template = jinja2.Environment(loader=jinja2.FileSystemLoader('')).get_template("pdf_template.html")

        by_year_table = []
        for year in list(self.stats1.keys())[:-1]:
            by_year_table.append([year, self.stats1[year], self.stats2[year], self.stats3[year], self.stats4[year]])

        for key in self.stats6:
            self.stats6[key] = round(self.stats6[key] * 100, 2)

        pdf_template = template.render({
            'name': self.vacancy_name,
            'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), self.img_filename),
            'stats': by_year_table,
            'stats5': self.stats5,
            'stats6': self.stats6
        })

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, filename, {"enable-local-file-access": ""},
                           configuration=config)


if __name__ == '__main__': InputConnect()
